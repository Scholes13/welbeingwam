import openpyxl

wb = openpyxl.load_workbook(r'I:\Project\Welbeing\Excel\Wellbeing Point.xlsx')

for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    ws = wb[sheet_name]
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=False):
        row_data = []
        for cell in row:
            val = cell.value
            if val is not None:
                row_data.append(f"{cell.coordinate}: {val}")
        if row_data:
            print(" | ".join(row_data))
